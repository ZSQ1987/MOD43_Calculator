import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
from datetime import datetime
import os
import configparser

# 尝试导入openpyxl库
try:
    from openpyxl import Workbook
except ImportError:
    print("缺少openpyxl库，请安装：pip install openpyxl")

# 定义 mod43_check_digit 函数
def mod43_check_digit(input_str):
    # 定义字符映射表
    scale_string = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ-. $/+%"
    # 初始化求和值
    total = 0
    # 遍历输入字符串每个字符
    for char in input_str:
        if char not in scale_string:
            raise ValueError(f"字符 {char} 不在映射表中")
        # 累加字符对应的索引
        total += scale_string.index(char)
    # 计算余数
    remainder = total % 43
    # 返回校验位
    return scale_string[remainder]

class MOD43_GUI:
    def __init__(self, root):
        self.root = root
        self.root.title("MOD43 校验位计算器")
        
        # 设置窗口大小
        window_width = 800
        window_height = 780
        
        # 计算屏幕居中位置
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # 设置窗口位置和大小
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # 设置图标
        self._set_window_icon()
        
        # 初始化配置文件路径
        self.config_file = os.path.join(os.getcwd(), "config.ini")
        # 读取配置
        self.string_length = self._load_config()
        # 初始化 GUI
        self._init_gui()
    
    def _set_window_icon(self):
        """设置窗口图标"""
        # 尝试设置图标
        import sys
        # 确定图标路径
        if getattr(sys, 'frozen', False):
            # 打包后的环境
            icon_path = os.path.join(sys._MEIPASS, "app_ico.ico")
        else:
            # 开发环境
            icon_path = os.path.join(os.getcwd(), "app_ico.ico")
        
        if os.path.exists(icon_path):
            try:
                self.root.iconbitmap(icon_path)
            except Exception as e:
                # 图标设置失败，静默处理
                pass
    
    def _load_config(self):
        """从配置文件加载配置"""
        config = configparser.ConfigParser()
        if os.path.exists(self.config_file):
            config.read(self.config_file)
            if "Settings" in config and "string_length" in config["Settings"]:
                try:
                    return int(config["Settings"]["string_length"])
                except (ValueError, TypeError):
                    pass
        # 默认值
        return 14
    
    def _save_config(self, length):
        """保存配置到配置文件"""
        config = configparser.ConfigParser()
        config["Settings"] = {"string_length": str(length)}
        with open(self.config_file, "w") as f:
            config.write(f)
    
    def _init_gui(self):
        """初始化 TTK 界面组件"""
        # 1. 顶部标题
        title_frame = ttk.Frame(self.root)
        title_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(title_frame, text="MOD43 校验位计算器", font=("Arial", 16, "bold")).pack(anchor=tk.CENTER)
        
        # 2. 输入区域
        input_frame = ttk.Frame(self.root)
        input_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(input_frame, text="输入字符串：").pack(side=tk.LEFT, padx=5)
        self.input_var = tk.StringVar(value="209 4YJ008Y8BK")
        self.input_entry = ttk.Entry(input_frame, textvariable=self.input_var, width=50)
        self.input_entry.pack(side=tk.LEFT, padx=5)
        
        # 添加点击事件，清空默认值
        def on_entry_click(event):
            if self.input_var.get() == "209 4YJ008Y8BK":
                self.input_var.set("")
        
        self.input_entry.bind('<Button-1>', on_entry_click)
        
        ttk.Button(input_frame, text="计算", command=self.calculate).pack(side=tk.LEFT, padx=5)
        ttk.Button(input_frame, text="导出Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(input_frame, text="配置", command=self.show_config).pack(side=tk.LEFT, padx=5)
        
        # 3. 结果显示区域
        result_frame = ttk.Frame(self.root)
        result_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(result_frame, text="计算结果：").pack(side=tk.LEFT, padx=5)
        self.result_var = tk.StringVar()
        ttk.Entry(result_frame, textvariable=self.result_var, width=10, state="readonly").pack(side=tk.LEFT, padx=5)
        
        # 显示对应的MOD32数值
        ttk.Label(result_frame, text="对应数值：").pack(side=tk.LEFT, padx=5)
        self.result_value_var = tk.StringVar()
        ttk.Entry(result_frame, textvariable=self.result_value_var, width=10, state="readonly").pack(side=tk.LEFT, padx=5)
        
        # 4. MOD43 字符映射表
        mapping_frame = ttk.LabelFrame(self.root, text="MOD43 字符映射表")
        mapping_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 定义字符映射表
        scale_string = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ-. $/+%"
        
        # 创建一个框架来容纳字符映射表
        table_frame = ttk.Frame(mapping_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 保存所有标签的引用，以便后续高亮
        self.map_labels = []
        
        # 创建表头
        for i in range(5):
            ttk.Label(table_frame, text="数值", font=('Arial', 10, 'bold'), borderwidth=1, relief=tk.RIDGE).grid(row=0, column=i*2, sticky='nsew', padx=1, pady=1)
            ttk.Label(table_frame, text="字符", font=('Arial', 10, 'bold'), borderwidth=1, relief=tk.RIDGE).grid(row=0, column=i*2+1, sticky='nsew', padx=1, pady=1)
        
        # 填充数据
        for row_idx in range(9):  # 9行
            row_labels = []
            for col_idx in range(5):  # 5列
                idx = row_idx + col_idx * 9
                if idx < len(scale_string):
                    # 创建数值标签
                    num_label = ttk.Label(table_frame, text=str(idx), borderwidth=1, relief=tk.RIDGE)
                    num_label.grid(row=row_idx+1, column=col_idx*2, sticky='nsew', padx=1, pady=1)
                    
                    # 创建字符标签
                    char = scale_string[idx]
                    if char == ' ':
                        char = '(空格)'
                    char_label = ttk.Label(table_frame, text=char, borderwidth=1, relief=tk.RIDGE)
                    char_label.grid(row=row_idx+1, column=col_idx*2+1, sticky='nsew', padx=1, pady=1)
                    
                    row_labels.append((num_label, char_label, idx))
                else:
                    # 空值
                    num_label = ttk.Label(table_frame, text='-', borderwidth=1, relief=tk.RIDGE)
                    num_label.grid(row=row_idx+1, column=col_idx*2, sticky='nsew', padx=1, pady=1)
                    
                    char_label = ttk.Label(table_frame, text='-', borderwidth=1, relief=tk.RIDGE)
                    char_label.grid(row=row_idx+1, column=col_idx*2+1, sticky='nsew', padx=1, pady=1)
                    
                    row_labels.append((num_label, char_label, -1))
            self.map_labels.append(row_labels)
        
        # 调整列宽
        for i in range(10):
            table_frame.grid_columnconfigure(i, weight=1, minsize=50)
        
        # 调整行高
        for i in range(10):
            table_frame.grid_rowconfigure(i, weight=1, minsize=25)
        
        # 5. 计算公式
        self.formula_frame = ttk.LabelFrame(self.root, text="计算公式")
        self.formula_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.formula_text = scrolledtext.ScrolledText(self.formula_frame, wrap=tk.WORD, font=("Consolas", 10), height=15)
        self.formula_text.pack(fill=tk.BOTH, expand=True)
        self.formula_text.config(state=tk.NORMAL)
        
        # 写入计算公式
        self.formula_text.insert(tk.END, "MOD43 校验位计算步骤：\n")
        self.formula_text.insert(tk.END, "1. 定义字符映射表：0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ-. $/+%\n")
        self.formula_text.insert(tk.END, "2. 遍历输入字符串的每个字符，获取其在映射表中的索引\n")
        self.formula_text.insert(tk.END, "3. 累加所有字符的索引值\n")
        self.formula_text.insert(tk.END, "4. 将累加和除以 43，取余数\n")
        self.formula_text.insert(tk.END, "5. 余数对应的映射表字符即为校验位\n\n")
        self.formula_text.insert(tk.END, "公式：校验位 = scale_string[sum(scale_string.index(char) for char in input_str) % 43]\n")
        
        self.formula_text.config(state=tk.DISABLED)
    
    def calculate(self):
        """计算 MOD43 校验位"""
        input_str = self.input_var.get().strip()
        if not input_str:
            self.center_messagebox("错误", "请输入字符串", 'error')
            return
        
        # 强制校验配置的字符串长度
        if len(input_str) != self.string_length:
            self.center_messagebox("错误", f"请输入{self.string_length}位字符串", 'error')
            return
        
        try:
            # 计算校验位并显示详细过程
            scale_string = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ-. $/+%"
            total = 0
            calculation_steps = []
            
            # 计算每个字符的索引和总和
            for i, char in enumerate(input_str):
                if char not in scale_string:
                    raise ValueError(f"字符 {char} 不在映射表中")
                char_index = scale_string.index(char)
                total += char_index
                calculation_steps.append(f"第{i+1}位 '{char}' → {char_index}")
            
            # 计算余数
            remainder = total % 43
            check_digit = scale_string[remainder]
            
            # 保存计算过程变量为实例变量，用于导出Excel
            self.calculation_steps = calculation_steps
            self.total = total
            self.remainder = remainder
            
            # 显示结果
            self.result_var.set(check_digit)
            
            # 显示对应的MOD32数值
            self.result_value_var.set(str(remainder))
            
            # 保存计算结果的余数，用于高亮显示
            self.result_remainder = remainder
            
            # 更新计算公式区域，显示详细计算过程
            self.formula_text.config(state=tk.NORMAL)
            self.formula_text.delete(1.0, tk.END)
            
            # 写入计算公式和详细过程
            self.formula_text.insert(tk.END, "MOD43 校验位计算步骤：\n")
            self.formula_text.insert(tk.END, "1. 定义字符映射表：0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ-. $/+%\n")
            self.formula_text.insert(tk.END, "2. 遍历输入字符串的每个字符，获取其在映射表中的索引\n")
            for step in calculation_steps:
                self.formula_text.insert(tk.END, f"   {step}\n")
            self.formula_text.insert(tk.END, f"3. 累加所有字符的索引值：{total}\n")
            self.formula_text.insert(tk.END, f"4. 将累加和除以 43，取余数：{total} ÷ 43 = {total // 43} 余 {remainder}\n")
            self.formula_text.insert(tk.END, f"5. 余数对应的映射表字符即为校验位：{remainder} → '{check_digit}'\n\n")
            
            self.formula_text.insert(tk.END, f"公式：校验位 = scale_string[{total} % 43] = scale_string[{remainder}] = '{check_digit}'\n")
            
            self.formula_text.config(state=tk.DISABLED)
            
            # 高亮显示字符映射表中对应的结果
            self.highlight_result_in_tree()
            
        except ValueError as e:
            self.center_messagebox("错误", str(e), 'error')
    
    def show_config(self):
        """显示配置弹窗，允许用户输入字符串长度"""
        # 创建配置窗口
        config_window = tk.Toplevel(self.root)
        config_window.title("配置")
        config_window.transient(self.root)  # 设置为主窗口的临时窗口
        config_window.grab_set()  # 模态窗口
        
        # 设置窗口大小
        window_width = 300
        window_height = 150
        
        # 计算居中位置
        screen_width = config_window.winfo_screenwidth()
        screen_height = config_window.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # 设置窗口位置
        config_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # 创建框架
        frame = ttk.Frame(config_window, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # 添加标签和输入框
        ttk.Label(frame, text="字符串长度：").grid(row=0, column=0, padx=10, pady=10, sticky=tk.E)
        
        # 创建字符串变量，用于存储输入的长度
        length_var = tk.StringVar(value=str(self.string_length))
        length_entry = ttk.Entry(frame, textvariable=length_var, width=10)
        length_entry.grid(row=0, column=1, padx=10, pady=10, sticky=tk.W)
        
        # 添加点击事件，清空默认值
        def on_entry_click(event):
            if length_var.get() == str(self.string_length):
                length_var.set("")
        
        length_entry.bind('<Button-1>', on_entry_click)
        
        # 聚焦到输入框
        length_entry.focus()
        
        # 确定按钮回调
        def on_ok():
            try:
                # 尝试将输入转换为整数
                new_length = int(length_var.get())
                if new_length > 0:
                    # 更新字符串长度配置
                    self.string_length = new_length
                    # 保存配置到文件
                    self._save_config(new_length)
                    config_window.destroy()
                else:
                    # 显示错误消息
                    self.center_messagebox("错误", "字符串长度必须大于0", 'error')
            except ValueError:
                # 显示错误消息
                self.center_messagebox("错误", "请输入有效的整数", 'error')
        
        # 取消按钮回调
        def on_cancel():
            config_window.destroy()
        
        # 添加按钮
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=1, column=0, columnspan=2, pady=20)
        
        ttk.Button(button_frame, text="确定", command=on_ok, width=10).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="取消", command=on_cancel, width=10).pack(side=tk.LEFT, padx=10)
    
    def highlight_result_in_tree(self):
        """在字符映射表中高亮显示计算结果对应的数值和字符"""
        if hasattr(self, 'result_remainder'):
            if hasattr(self, 'map_labels'):
                # 取消之前的高亮
                for row in self.map_labels:
                    for num_label, char_label, idx in row:
                        # 恢复默认背景色
                        num_label.config(background='')
                        char_label.config(background='')
                
                # 查找并高亮显示结果对应的数值和字符
                for row in self.map_labels:
                    for num_label, char_label, idx in row:
                        if idx == self.result_remainder:
                            # 设置绿色背景
                            num_label.config(background='#00FF00')
                            char_label.config(background='#00FF00')
                            break
    
    def center_messagebox(self, title, message, icon='info', filepath=None):
        """创建居中显示的消息框"""
        # 创建一个自定义消息框窗口
        msg_window = tk.Toplevel(self.root)
        msg_window.title(title)
        msg_window.transient(self.root)  # 设置为主窗口的临时窗口
        msg_window.grab_set()  # 模态窗口
        
        # 设置窗口大小
        window_width = 500
        # 计算消息需要的高度
        lines = message.count('\n') + 1
        window_height = max(150, min(400, 50 + lines * 20))
        
        # 计算居中位置
        screen_width = msg_window.winfo_screenwidth()
        screen_height = msg_window.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # 设置窗口位置
        msg_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # 添加消息文本
        message_label = tk.Label(msg_window, text=message, wraplength=window_width-40, justify=tk.LEFT, font=('Arial', 10))
        
        # 添加确定按钮
        def on_ok():
            msg_window.destroy()
        
        # 导入os模块用于打开文件
        import os
        
        # 添加打开文件按钮
        def on_open_file():
            if filepath:
                try:
                    os.startfile(filepath)
                except Exception as e:
                    print(f"打开文件失败: {e}")
            msg_window.destroy()
        
        # 使用网格布局
        msg_window.grid_columnconfigure(0, weight=1)
        msg_window.grid_rowconfigure(0, weight=1)
        msg_window.grid_rowconfigure(1, weight=0)
        
        message_label.grid(row=0, column=0, padx=20, pady=20, sticky='nsew')
        
        # 如果提供了文件路径，显示两个按钮
        if filepath:
            button_frame = tk.Frame(msg_window)
            button_frame.grid(row=1, column=0, padx=20, pady=10, sticky='se')
            
            open_button = tk.Button(button_frame, text="打开Excel", command=on_open_file, width=10)
            open_button.pack(side=tk.LEFT, padx=5)
            
            ok_button = tk.Button(button_frame, text="确定", command=on_ok, width=10)
            ok_button.pack(side=tk.LEFT, padx=5)
        else:
            # 否则只显示确定按钮
            ok_button = tk.Button(msg_window, text="确定", command=on_ok, width=10)
            ok_button.grid(row=1, column=0, padx=20, pady=10, sticky='se')
        
        # 等待用户点击
        msg_window.wait_window(msg_window)

    def export_to_excel(self):
        """将计算结果导出到Excel文件"""
        input_str = self.input_var.get().strip()
        check_digit = self.result_var.get()
        
        if not input_str or check_digit in ["请输入字符串", "错误：请输入14位字符串", "错误：字符不在映射表中"]:
            self.center_messagebox("错误", "请先计算校验位", 'error')
            return
        
        try:
            # 检查是否安装了openpyxl
            from openpyxl import Workbook
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 创建第一个sheet：MOD32字符集
            ws1 = wb.active
            ws1.title = "MOD32字符集"
            
            # 写入MOD43字符映射表
            ws1['A1'] = "MOD43字符映射表"
            
            # 设置映射表表头
            headers = ['数值', '字符', '数值', '字符', '数值', '字符', '数值', '字符', '数值', '字符']
            for i, header in enumerate(headers, 1):
                ws1.cell(row=2, column=i, value=header)
            
            # 填充映射表数据，按照用户提供的样式
            scale_string = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ-. $/+%"
            col_headers = [0, 9, 18, 27, 36]
            
            # 获取计算结果对应的数值
            result_value = None
            if hasattr(self, 'remainder'):
                result_value = self.remainder
            
            # 导入样式模块
            from openpyxl.styles import PatternFill
            
            # 定义绿色背景
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            
            for row_idx in range(8):  # 8行
                for col_idx in range(5):  # 5列组
                    idx = row_idx + col_idx * 9
                    if idx < len(scale_string):
                        # 数值
                        num_cell = ws1.cell(row=row_idx+3, column=col_idx*2+1, value=idx)
                        # 字符
                        char = scale_string[idx]
                        char_cell = ws1.cell(row=row_idx+3, column=col_idx*2+2, value=char if char != ' ' else '(空格)')
                        
                        # 如果是计算结果对应的数值和字符，设置绿色背景
                        if result_value is not None and idx == result_value:
                            num_cell.fill = green_fill
                            char_cell.fill = green_fill
                    else:
                        # 空值
                        ws1.cell(row=row_idx+3, column=col_idx*2+1, value='-')
                        ws1.cell(row=row_idx+3, column=col_idx*2+2, value='-')
            
            # 调整列宽
            ws1.column_dimensions['A'].width = 10
            ws1.column_dimensions['B'].width = 10
            ws1.column_dimensions['C'].width = 10
            ws1.column_dimensions['D'].width = 10
            ws1.column_dimensions['E'].width = 10
            ws1.column_dimensions['F'].width = 10
            ws1.column_dimensions['G'].width = 10
            ws1.column_dimensions['H'].width = 10
            ws1.column_dimensions['I'].width = 10
            ws1.column_dimensions['J'].width = 10
            
            # 创建第二个sheet：计算结果
            ws2 = wb.create_sheet(title="计算结果")
            
            # 写入基本信息
            ws2['A1'] = "基本信息"
            ws2['A2'] = "输入字符串"
            ws2['B2'] = input_str
            ws2['A3'] = "校验位"
            ws2['B3'] = check_digit
            ws2['A4'] = "计算时间"
            ws2['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 写入计算公式部分
            ws2['A6'] = "计算公式"
            ws2['A7'] = "MOD43 校验位计算步骤："
            
            # 步骤1：定义字符映射表
            ws2['A8'] = "1. 定义字符映射表：0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ-. $/+%"
            
            # 步骤2：遍历输入字符串的每个字符
            ws2['A9'] = "2. 遍历输入字符串的每个字符，获取其在映射表中的索引"
            
            if hasattr(self, 'calculation_steps'):
                row = 10
                for step in self.calculation_steps:
                    ws2['A' + str(row)] = step
                    row += 1
            
            # 步骤3：累加所有字符的索引值
            if hasattr(self, 'total'):
                ws2['A' + str(row)] = f"3. 累加所有字符的索引值：{self.total}"
                row += 1
            
            # 步骤4：将累加和除以43，取余数
            if hasattr(self, 'total') and hasattr(self, 'remainder'):
                ws2['A' + str(row)] = f"4. 将累加和除以 43，取余数：{self.total} ÷ 43 = {self.total // 43} 余 {self.remainder}"
                row += 1
            
            # 步骤5：余数对应的映射表字符即为校验位
            if hasattr(self, 'remainder'):
                ws2['A' + str(row)] = f"5. 余数对应的映射表字符即为校验位：{self.remainder} → '{check_digit}'"
                row += 1
            
            # 公式
            if hasattr(self, 'total') and hasattr(self, 'remainder'):
                ws2['A' + str(row)] = f"公式：校验位 = scale_string[{self.total} % 43] = scale_string[{self.remainder}] = '{check_digit}'"
            
            # 调整列宽
            ws2.column_dimensions['A'].width = 50
            ws2.column_dimensions['B'].width = 30
            
            # 保存文件
            filename = f"MOD43_Result_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            filepath = os.path.join(os.getcwd(), filename)
            wb.save(filepath)
            
            self.center_messagebox("成功", f"导出成功，文件保存到：{filepath}", 'info', filepath)
            
        except ImportError:
            self.center_messagebox("错误", "缺少openpyxl库，请安装：pip install openpyxl", 'error')
        except Exception as e:
            self.center_messagebox("错误", f"导出失败：{str(e)}", 'error')

if __name__ == "__main__":
    root = tk.Tk()
    app = MOD43_GUI(root)
    root.mainloop()
