import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import socket
import threading
import time
from datetime import datetime
import os

class IPLPrinterSimulator:
    def __init__(self, root):
        self.root = root
        self.root.title("IPL标签打印机模拟器 (端口9100)")
        
        # 设置窗口大小
        window_width = 800
        window_height = 600
        
        # 计算屏幕居中位置
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # 设置窗口位置和大小
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # 核心变量
        self.HOST = '127.0.0.1'  # 只监听本地连接
        self.PORT = 9100
        self.server_socket = None
        self.is_listening = False  # 监听状态标记
        self.client_sockets = []   # 保存客户端连接
        self.lock = threading.Lock()  # 线程锁，保护客户端列表
        self.print_count = 0  # 打印次数计数器

        # 初始化GUI
        self._init_gui()
        
        # 确保窗口显示前完成所有初始化，减少闪烁
        self.root.update_idletasks()
        self.root.deiconify()

    def _init_gui(self):
        """初始化TTK界面组件"""
        # 1. 顶部状态栏
        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(status_frame, text="监听状态：").grid(row=0, column=0, padx=5)
        self.status_label = ttk.Label(status_frame, text="未监听", foreground="red")
        self.status_label.grid(row=0, column=1, padx=5)

        ttk.Label(status_frame, text="当前连接客户端：").grid(row=0, column=2, padx=5)
        self.client_count_label = ttk.Label(status_frame, text="0")
        self.client_count_label.grid(row=0, column=3, padx=5)

        # 2. 功能按钮区
        btn_frame = ttk.Frame(self.root)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)

        self.start_btn = ttk.Button(btn_frame, text="启动监听", command=self.start_listening)
        self.start_btn.pack(side=tk.LEFT, padx=5)

        self.stop_btn = ttk.Button(btn_frame, text="停止监听", command=self.stop_listening, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        ttk.Button(btn_frame, text="清空日志", command=self.clear_log).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="保存日志", command=self.save_commands).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="解析app.log", command=self.parse_app_log).pack(side=tk.LEFT, padx=5)

        # 3. 日志/指令显示区（带滚动条）
        log_frame = ttk.Frame(self.root)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        ttk.Label(log_frame, text="IPL指令/日志记录：").pack(anchor=tk.W)
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, font=("Consolas", 10))
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.config(state=tk.DISABLED)  # 初始设为只读

    def update_log(self, content, is_ipl_command=False):
        """更新日志文本框（线程安全）"""
        # 区分普通日志、IPL指令、ZPL指令和XML内容
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
        if is_ipl_command:
            log_content = f"{timestamp}  {content}\n"
            tag = "ipl"
        elif content.startswith('^') or 'ZPL' in content:
            log_content = f"{timestamp}  {content}\n"
            tag = "zpl"
        elif '<xpml>' in content or '</xpml>' in content:
            log_content = f"{timestamp}  {content}\n"
            tag = "xml"
        else:
            log_content = f"{timestamp}  {content}\n"
            tag = "normal"

        # 用after方法确保在GUI主线程更新
        self.root.after(0, self._do_update_log, log_content, tag)

    def _do_update_log(self, content, tag):
        """实际更新日志的方法（在GUI主线程执行）"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, content, tag)
        self.log_text.tag_config("ipl", foreground="darkgreen")  # IPL指令标绿
        self.log_text.tag_config("zpl", foreground="darkblue")  # ZPL指令标蓝
        self.log_text.tag_config("xml", foreground="purple")  # XML内容标紫
        self.log_text.tag_config("normal", foreground="black")
        self.log_text.see(tk.END)  # 自动滚动到末尾
        self.log_text.config(state=tk.DISABLED)

    def clear_log(self):
        """清空日志"""
        # 重置打印计数
        with self.lock:
            self.print_count = 0
        # 更新客户端计数标签
        self.client_count_label.config(text="0")
        # 清空日志
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
        # 按照app.log格式记录清空信息
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
        clear_msg = f"{timestamp}  日志已清空"
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, clear_msg + '\n', "normal")
        self.log_text.config(state=tk.DISABLED)

    def save_commands(self):
        """保存当前窗口输出的日志到文件"""
        log_content = self.log_text.get(1.0, tk.END)
        if not log_content.strip():
            self.center_messagebox("提示", "日志为空，无需保存", 'warning')
            return

        # 使用YYYY-MM-DD.log格式的文件名
        file_path = f"{datetime.now().strftime('%Y-%m-%d')}.log"
        try:
            # 保存所有日志内容
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(log_content)
            self.center_messagebox("成功", f"日志已保存到：{os.path.abspath(file_path)}", 'info')
            # 记录保存信息
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
            save_msg = f"{timestamp}  日志已保存到 {file_path}"
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, save_msg + '\n', "normal")
            self.log_text.config(state=tk.DISABLED)
        except Exception as e:
            self.center_messagebox("错误", f"保存失败：{str(e)}", 'error')

    def handle_client(self, client_socket, client_address):
        """处理单个客户端连接（接收IPL指令）"""
        # 更新客户端计数
        with self.lock:
            self.client_sockets.append(client_socket)
            # 显示打印总数而不是客户端数量
            self.root.after(0, lambda: self.client_count_label.config(text=str(self.print_count)))

        try:
            # 循环接收客户端数据（模拟打印机持续接收指令）
            while self.is_listening:
                data = client_socket.recv(4096)  # 缓冲区4096字节
                if not data:
                    break  # 客户端关闭连接

                # 增加打印计数
                with self.lock:
                    self.print_count += 1
                    # 更新客户端计数标签，显示打印总数
                    self.root.after(0, lambda: self.client_count_label.config(text=str(self.print_count)))
                # 只记录打印次数
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
                print_msg = f"{timestamp}  {self.print_count} ok"
                self.log_text.config(state=tk.NORMAL)
                self.log_text.insert(tk.END, print_msg + '\n', "normal")
                self.log_text.config(state=tk.DISABLED)

                # 模拟打印机返回确认（可选，部分客户端需要确认）
                ack_response = b"\x06"  # ASCII ACK（确认）
                client_socket.sendall(ack_response)

        except Exception as e:
            # 忽略正常的连接关闭错误
            if "10053" not in str(e) and "10054" not in str(e):
                # 错误信息也按照app.log格式保存
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
                error_msg = f"{timestamp}  客户端通信错误 {client_address}：{str(e)}"
                self.log_text.config(state=tk.NORMAL)
                self.log_text.insert(tk.END, error_msg + '\n', "normal")
                self.log_text.config(state=tk.DISABLED)
        finally:
            # 关闭客户端连接，更新计数
            client_socket.close()
            with self.lock:
                if client_socket in self.client_sockets:
                    self.client_sockets.remove(client_socket)
                self.root.after(0, lambda: self.client_count_label.config(text=str(len(self.client_sockets))))
            # 不再记录客户端断开信息

    def start_listening(self):
        """启动TCP监听"""
        try:
            # 检查端口是否被占用
            import socket
            test_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            test_socket.settimeout(1)
            try:
                test_socket.bind((self.HOST, self.PORT))
                test_socket.close()
            except socket.error as e:
                test_socket.close()
                # 端口被占用，提供选项
                from tkinter import simpledialog
                new_port = simpledialog.askinteger("端口被占用", f"端口 {self.PORT} 已被占用，请输入其他端口：", minvalue=1024, maxvalue=65535, initialvalue=9201)
                if new_port is None:
                    return  # 用户取消
                self.PORT = new_port
                # 更新窗口标题
                self.root.title(f"IPL标签打印机模拟器 (端口{self.PORT})")

            # 创建TCP服务器套接字
            self.server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.server_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)  # 端口复用
            self.server_socket.bind((self.HOST, self.PORT))
            self.server_socket.listen(5)  # 最大挂起连接数5
            self.is_listening = True

            # 更新GUI状态
            self.status_label.config(text=f"已监听 {self.HOST}:{self.PORT}", foreground="green")
            self.start_btn.config(state=tk.DISABLED)
            self.stop_btn.config(state=tk.NORMAL)
            # 按照app.log格式记录启动信息
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
            start_msg = f"{timestamp}  开始监听 {self.HOST}:{self.PORT} 端口（IPL打印机模拟器）"
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, start_msg + '\n', "normal")
            self.log_text.config(state=tk.DISABLED)

            # 启动监听线程（避免阻塞GUI）
            listen_thread = threading.Thread(target=self.accept_clients, daemon=True)
            listen_thread.start()

        except Exception as e:
            self.center_messagebox("错误", f"启动监听失败：{str(e)}", 'error')
            self.is_listening = False
            self.status_label.config(text="启动失败", foreground="red")

    def accept_clients(self):
        """循环接收客户端连接"""
        while self.is_listening:
            try:
                client_socket, client_address = self.server_socket.accept()
                # 启动线程处理单个客户端
                client_thread = threading.Thread(target=self.handle_client, args=(client_socket, client_address), daemon=True)
                client_thread.start()
            except Exception as e:
                if self.is_listening:  # 非主动停止时记录错误
                    self.update_log(f"接收连接错误：{str(e)}")
                break

    def stop_listening(self):
        """停止监听，关闭所有连接"""
        self.is_listening = False
        # 关闭服务器套接字
        if self.server_socket:
            try:
                self.server_socket.close()
            except:
                pass
        # 关闭所有客户端连接
        with self.lock:
            for sock in self.client_sockets:
                try:
                    sock.close()
                except:
                    pass
            self.client_sockets.clear()

        # 重置打印计数
        with self.lock:
            self.print_count = 0
        # 更新GUI状态
        self.status_label.config(text="未监听", foreground="red")
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.client_count_label.config(text="0")
        # 按照app.log格式记录停止信息
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
        stop_msg = f"{timestamp}  已停止监听，所有连接已关闭"
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, stop_msg + '\n', "normal")
        self.log_text.config(state=tk.DISABLED)

    def parse_app_log(self):
        """解析app.log文件，提取时间和二维码信息，保存为Excel格式"""
        import os
        import re
        try:
            # 导入openpyxl库，用于创建Excel文件
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill
            from openpyxl.worksheet.filters import AutoFilter
        except ImportError:
            self.center_messagebox("错误", "缺少openpyxl库，请先安装：pip install openpyxl", 'error')
            return

        # app.log文件路径
        app_log_path = r"D:\ZF_CD_FGLable\FGLABEL(1)\Log\app.log"
        
        # 检查文件是否存在
        if not os.path.exists(app_log_path):
            self.center_messagebox("错误", f"文件不存在：{app_log_path}", 'error')
            return

        try:
            # 读取app.log文件，忽略编码错误
            with open(app_log_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()

            # 记录文件读取信息
            timestamp_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
            read_msg = f"{timestamp_now}  读取app.log文件成功，共 {len(lines)} 行"
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, read_msg + '\n', "normal")
            self.log_text.config(state=tk.DISABLED)

            # 显示前5行内容，用于调试
            debug_msg = f"{timestamp_now}  前5行内容："
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, debug_msg + '\n', "normal")
            for i, line in enumerate(lines[:5]):
                debug_line = f"{timestamp_now}  第{i+1}行：{line.strip()}"
                self.log_text.insert(tk.END, debug_line + '\n', "normal")
            self.log_text.config(state=tk.DISABLED)

            # 按照新规则解析：以包含<xpml>的时间行为开头，以<xpml><end/></xpml>为结尾
            stages = []
            current_stage = []
            start_time = ""
            
            # 记录解析开始信息
            parse_start = f"{timestamp_now}  开始按新规则解析文件"
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, parse_start + '\n', "normal")
            self.log_text.config(state=tk.DISABLED)

            for i, line in enumerate(lines):
                line = line.strip()
                
                # 检查是否是阶段开始：包含时间格式 + <xpml>
                time_match = re.search(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', line)
                if time_match and '<xpml>' in line:
                    # 如果已经有正在处理的阶段，先处理完
                    if current_stage:
                        # 查找当前阶段中的二维码信息
                        qr_code = ""
                        for stage_line in current_stage:
                            qr_match = re.search(r'\*(.*?)\*', stage_line)
                            if qr_match:
                                qr_code = qr_match.group(1)
                                break
                        
                        if start_time and qr_code:
                            stages.append((start_time, qr_code))
                            # 记录解析到的信息
                            parse_info = f"{timestamp_now}  解析到：时间={start_time}, 二维码={qr_code}"
                            self.log_text.config(state=tk.NORMAL)
                            self.log_text.insert(tk.END, parse_info + '\n', "normal")
                            self.log_text.config(state=tk.DISABLED)
                    
                    # 开始新的阶段
                    current_stage = [line]
                    # 提取时间（已经在上面的time_match中获取）
                    start_time = time_match.group(0)
                    time_info = f"{timestamp_now}  提取时间成功：{start_time}"
                    self.log_text.config(state=tk.NORMAL)
                    self.log_text.insert(tk.END, time_info + '\n', "normal")
                    self.log_text.config(state=tk.DISABLED)
                
                # 检查是否是阶段继续
                elif current_stage:
                    current_stage.append(line)
                    
                    # 检查是否是阶段结束：<xpml><end/></xpml>
                    if line == '<xpml><end/></xpml>':
                        # 查找当前阶段中的二维码信息
                        qr_code = ""
                        for stage_line in current_stage:
                            qr_match = re.search(r'\*(.*?)\*', stage_line)
                            if qr_match:
                                qr_code = qr_match.group(1)
                                qr_info = f"{timestamp_now}  提取二维码成功：{qr_code}"
                                self.log_text.config(state=tk.NORMAL)
                                self.log_text.insert(tk.END, qr_info + '\n', "normal")
                                self.log_text.config(state=tk.DISABLED)
                                break
                        else:
                            qr_info = f"{timestamp_now}  提取二维码失败"
                            self.log_text.config(state=tk.NORMAL)
                            self.log_text.insert(tk.END, qr_info + '\n', "normal")
                            self.log_text.config(state=tk.DISABLED)
                        
                        if start_time and qr_code:
                            stages.append((start_time, qr_code))
                            # 记录解析到的信息
                            parse_info = f"{timestamp_now}  解析到：时间={start_time}, 二维码={qr_code}"
                            self.log_text.config(state=tk.NORMAL)
                            self.log_text.insert(tk.END, parse_info + '\n', "normal")
                            self.log_text.config(state=tk.DISABLED)
                        
                        # 重置当前阶段
                        current_stage = []
                        start_time = ""

            # 处理最后一个未结束的阶段
            if current_stage and start_time:
                # 查找二维码信息
                qr_code = ""
                for stage_line in current_stage:
                    qr_match = re.search(r'\*(.*?)\*', stage_line)
                    if qr_match:
                        qr_code = qr_match.group(1)
                        break
                
                if qr_code:
                    stages.append((start_time, qr_code))
                    # 记录解析到的信息
                    parse_info = f"{timestamp_now}  解析到：时间={start_time}, 二维码={qr_code}"
                    self.log_text.config(state=tk.NORMAL)
                    self.log_text.insert(tk.END, parse_info + '\n', "normal")
                    self.log_text.config(state=tk.DISABLED)

            # 记录阶段信息
            stage_msg = f"{timestamp_now}  共解析到 {len(stages)} 个阶段"
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, stage_msg + '\n', "normal")
            self.log_text.config(state=tk.DISABLED)

            if not stages:
                self.center_messagebox("提示", "未解析到有效数据", 'warning')
                return

            # 创建Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = "解析结果"

            # 设置表头
            ws['A1'] = "时间"
            ws['B1'] = "二维码"
            ws['C1'] = "字符长度"

            # 为第一行添加筛选功能
            ws.auto_filter.ref = ws.dimensions

            # 填充数据
            for row_idx, (time_str, qr_code) in enumerate(stages, start=2):
                ws[f'A{row_idx}'] = time_str
                ws[f'B{row_idx}'] = qr_code
                # 计算二维码字符长度
                length = len(qr_code)
                ws[f'C{row_idx}'] = length
                
                # 设置颜色格式
                cell = ws[f'C{row_idx}']
                if length == 15:
                    # 设置为绿色
                    cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                else:
                    # 设置为红色
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

            # 生成Excel文件名：年月日+00001序号
            date_str = datetime.now().strftime("%Y%m%d")
            
            # 查找当前目录中相同日期的文件，确定序号
            import glob
            existing_files = glob.glob(f"{date_str}*.xlsx")
            
            # 提取序号并找到最大序号
            max_seq = 0
            for file in existing_files:
                # 提取文件名中的序号部分
                import re
                match = re.search(rf'{date_str}(\d{{5}})\.xlsx', file)
                if match:
                    seq = int(match.group(1))
                    if seq > max_seq:
                        max_seq = seq
            
            # 生成新序号
            new_seq = max_seq + 1
            excel_filename = f"{date_str}{new_seq:05d}.xlsx"
            excel_path = os.path.join(os.getcwd(), excel_filename)
            wb.save(excel_path)

            self.center_messagebox("成功", f"解析完成，已保存到：{excel_path}", 'info')
            # 记录解析结果
            timestamp_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
            parse_msg = f"{timestamp_now}  解析app.log完成，保存为：{excel_filename}"
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, parse_msg + '\n', "normal")
            self.log_text.config(state=tk.DISABLED)

        except Exception as e:
            self.center_messagebox("错误", f"解析失败：{str(e)}", 'error')
            # 记录错误信息
            timestamp_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
            error_msg = f"{timestamp_now}  解析错误：{str(e)}"
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, error_msg + '\n', "normal")
            self.log_text.config(state=tk.DISABLED)

    def center_messagebox(self, title, message, icon='info'):
        """创建居中显示的消息框"""
        # 创建一个自定义消息框窗口
        msg_window = tk.Toplevel(self.root)
        msg_window.title(title)
        msg_window.transient(self.root)  # 设置为主窗口的临时窗口
        msg_window.grab_set()  # 模态窗口
        
        # 根据消息长度调整窗口大小
        window_width = 500
        # 计算消息需要的高度
        lines = message.count('\n') + 1
        window_height = max(150, min(400, 50 + lines * 20))
        
        # 计算居中位置
        screen_width = msg_window.winfo_screenwidth()
        screen_height = msg_window.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # 设置窗口位置
        msg_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # 添加消息文本
        message_label = tk.Label(msg_window, text=message, wraplength=window_width-40, justify=tk.LEFT, font=('Arial', 10))
        
        # 添加确定按钮
        def on_ok():
            msg_window.destroy()
        
        ok_button = tk.Button(msg_window, text="确定", command=on_ok, width=10)
        
        # 使用网格布局
        msg_window.grid_columnconfigure(0, weight=1)
        msg_window.grid_rowconfigure(0, weight=1)
        msg_window.grid_rowconfigure(1, weight=0)
        
        message_label.grid(row=0, column=0, padx=20, pady=20, sticky='nsew')
        ok_button.grid(row=1, column=0, padx=20, pady=10, sticky='se')
        
        # 等待用户点击
        msg_window.wait_window(msg_window)

    def on_closing(self):
        """窗口关闭时的清理"""
        if self.is_listening:
            self.stop_listening()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    # 隐藏窗口，等初始化完成后再显示
    root.withdraw()
    app = IPLPrinterSimulator(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)  # 关闭窗口时清理
    root.mainloop()